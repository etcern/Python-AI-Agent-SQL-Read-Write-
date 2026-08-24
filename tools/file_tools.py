"""File tools - let agents read and write files in the workspace.

Split into read-only and write groups so agents can be given granular access.
All paths are relative to WORKSPACE_DIR (defined in config.py).
Supports PDF, DOCX, XLSX, CSV, images (OCR), JSON, and plain text.
Ref: https://docs.python.org/3/library/os.path.html
Ref: https://openpyxl.readthedocs.io/
"""

import json
import os
from langchain_core.tools import tool
from logging_utils import log_panel
from config import WORKSPACE_DIR
from tools.extractors import extract_text


# --- Read-only tools ---

@tool(parse_docstring=True)
def read_file(file_path: str, reasoning: str = "") -> str:
    """Read the contents of a file.

    Supports text files, PDF, DOCX, XLSX, CSV, JSON, and images (OCR).
    Large files are automatically truncated.

    Args:
        file_path: Path to the file (relative to workspace folder).
        reasoning: Optional. Why you need to read this file.

    Returns:
        The file contents as text, or an error message.
    """
    if reasoning:
        log_panel(reasoning, title=f"read_file({file_path}) - Reasoning")
    full_path = os.path.join(WORKSPACE_DIR, file_path)
    if not os.path.exists(full_path):
        return f"File not found: {file_path}"
    content = extract_text(full_path)
    log_panel(content[:500], title=f"read_file({file_path}) - Result")
    return content


@tool(parse_docstring=True)
def list_files(directory: str = ".", reasoning: str = "") -> str:
    """List files in a directory.

    Args:
        directory: Directory path (relative to workspace folder). Defaults to root.
        reasoning: Optional. Why you need to list files.

    Returns:
        A list of file and folder names.
    """
    if reasoning:
        log_panel(reasoning, title=f"list_files({directory}) - Reasoning")
    full_path = os.path.join(WORKSPACE_DIR, directory)
    if not os.path.exists(full_path):
        return f"Directory not found: {directory}"
    entries = os.listdir(full_path)
    result = "\n".join(entries) if entries else "(empty directory)"
    log_panel(result, title=f"list_files({directory}) - Result")
    return result


# --- Write tools ---

@tool(parse_docstring=True)
def write_file(file_path: str, content: str, reasoning: str = "") -> str:
    """Write text content to a file. Creates the file if it doesn't exist.

    For spreadsheets (.xlsx), use create_spreadsheet instead.
    This tool writes plain text only.

    Args:
        file_path: Path to the file (relative to workspace folder).
        content: The text content to write.
        reasoning: Optional. Why you are writing this file.

    Returns:
        Confirmation message.
    """
    if reasoning:
        log_panel(reasoning, title=f"write_file({file_path}) - Reasoning")

    # -- Reject binary formats that need specialized tools --
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return (
            f"Cannot write {ext} files with write_file (it writes plain text). "
            "Use create_spreadsheet instead for Excel files."
        )

    full_path = os.path.join(WORKSPACE_DIR, file_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "w", encoding="utf-8") as f:
        f.write(content)
    abs_path = os.path.abspath(full_path)
    log_panel(f"Wrote {len(content)} characters to {abs_path}", title=f"write_file({file_path}) - Result")
    return f"File written: {file_path} ({len(content)} characters)\nSaved to: {abs_path}"


@tool(parse_docstring=True)
def create_spreadsheet(file_path: str, data: str, reasoning: str = "") -> str:
    """Create an Excel spreadsheet (.xlsx) with structured data.

    Pass data as a JSON string: a list of lists where the first list is headers.
    Example: '[["Name", "Age"], ["Alice", 30], ["Bob", 25]]'

    For multiple sheets, pass a JSON object with sheet names as keys:
    '{"Staff": [["Name", "Role"], ["Alice", "Dev"]], "Budget": [["Item", "Cost"], ["Server", 50]]}'

    Args:
        file_path: Path for the .xlsx file (relative to workspace folder).
        data: JSON string - list of rows, or object of {sheet_name: rows}.
        reasoning: Optional. Why you are creating this spreadsheet.

    Returns:
        Confirmation message or error if openpyxl is not installed.
    """
    if reasoning:
        log_panel(reasoning, title=f"create_spreadsheet({file_path}) - Reasoning")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl is not installed. Run: pip install openpyxl"

    # -- Parse the data --
    try:
        parsed = json.loads(data)
    except json.JSONDecodeError as e:
        return f"Invalid JSON data: {e}"

    # -- Normalize to {sheet_name: rows} format --
    if isinstance(parsed, list):
        sheets = {"Sheet1": parsed}
    elif isinstance(parsed, dict):
        sheets = parsed
    else:
        return "Data must be a JSON list of rows or object of {sheet: rows}"

    # -- Create workbook --
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    header_font_color = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    total_rows = 0
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])  # Excel limit
        if not isinstance(rows, list):
            continue

        for r_idx, row in enumerate(rows, start=1):
            if not isinstance(row, list):
                row = [row]
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                # -- Style header row --
                if r_idx == 1:
                    cell.font = header_font_color
                    cell.fill = header_fill
            total_rows += 1

        # -- Auto-fit column widths (approximate) --
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # -- Ensure .xlsx extension --
    if not file_path.lower().endswith(".xlsx"):
        file_path += ".xlsx"

    full_path = os.path.join(WORKSPACE_DIR, file_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    wb.save(full_path)

    abs_path = os.path.abspath(full_path)
    sheet_count = len(sheets)
    log_panel(
        f"Created {file_path}: {sheet_count} sheet(s), {total_rows} rows at {abs_path}",
        title=f"create_spreadsheet({file_path}) - Result",
    )
    return (
        f"Spreadsheet created: {file_path} ({sheet_count} sheet(s), {total_rows} rows)\n"
        f"Saved to: {abs_path}\n"
        f"IMPORTANT: Tell the user the exact file path so they can find it."
    )


# --- Tool group accessors ---

def get_file_read_tools() -> list:
    """Read-only file tools - safe to give any agent."""
    return [read_file, list_files]


def get_file_write_tools() -> list:
    """Write tools - for agents that need disk access."""
    return [write_file, create_spreadsheet]


def get_file_tools() -> list:
    """All file tools (read + write). Backward compatible."""
    return [read_file, write_file, create_spreadsheet, list_files]
